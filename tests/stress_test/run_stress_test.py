"""
Stress Test Runner
==================
Reads every question from 'RAG STRESS TEST 2.xlsx' (column: "questions"),
pipes each one through the GraphRAGPipeline (same logic as main.py),
and writes the model answer into a new column "model_answer" in the
same workbook.

Usage (from the project root):
    python tests/StressTest/run_stress_test.py

Optional flags:
    --excel   Path to the Excel file  (default: auto-detected)
    --sheet   Sheet name              (default: first sheet)
    --delay   Seconds to wait between queries to avoid rate-limits (default: 2)
"""

import argparse
import sys
import time
from pathlib import Path

# Make sure the project root is on sys.path so 'graphrag' is importable
PROJECT_ROOT = Path(__file__).resolve().parents[2]   # tests/StressTest/../../
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_excel(folder: Path) -> Path:
    """Return the first .xlsx file found in *folder*."""
    files = list(folder.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("No .xlsx file found in " + str(folder))
    return files[0]


def find_question_col(ws) -> int:
    """Return the 1-based column index for the question column.

    Matches (case-insensitive):
      - 'questions'
      - 'user question'
      - any header containing the word 'question'
    """
    for cell in ws[1]:
        if cell.value and "question" in str(cell.value).strip().lower():
            return cell.column
    headers = [c.value for c in ws[1]]
    raise ValueError(
        "Could not find a question column (expected a header containing 'question'). "
        "Headers found: " + str(headers)
    )


def find_or_create_answer_col(ws, question_col: int) -> int:
    """Return (or create) the 'model_answer' column."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == "model_answer":
            return cell.column
    answer_col = ws.max_column + 1
    ws.cell(row=1, column=answer_col, value="model_answer")
    return answer_col


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------

def run_stress_test(excel_path: Path, sheet_name, delay: float):
    print("\n[INFO] Loading workbook: " + str(excel_path))
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    print("       Sheet: '" + ws.title + "'  |  Rows: " + str(ws.max_row) + "  |  Cols: " + str(ws.max_column))

    q_col = find_question_col(ws)
    a_col = find_or_create_answer_col(ws, q_col)
    print("       Question column: " + str(q_col) + "  |  Answer column: " + str(a_col) + "\n")

    # Collect questions (skip header row 1)
    questions = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=q_col).value
        if val and str(val).strip():
            questions.append((row, str(val).strip()))

    if not questions:
        print("[WARN] No questions found. Exiting.")
        return

    print("[INFO] Found " + str(len(questions)) + " question(s) to process.\n")
    print("=" * 72)

    # Initialise pipeline once
    print("[INIT] Initialising GraphRAG pipeline ...")
    from graphrag.pipeline.graphrag_pipeline import GraphRAGPipeline
    try:
        pipeline = GraphRAGPipeline()
    except Exception as e:
        print("\n[FAIL] Pipeline init failed: " + str(e))
        sys.exit(1)
    print("[OK]   Pipeline ready.\n")
    print("=" * 72)

    total = len(questions)
    for idx, (row, question) in enumerate(questions, start=1):
        print("\n[" + str(idx) + "/" + str(total) + "] Row " + str(row))
        print("   Q: " + question)

        # Skip if already answered (re-run protection)
        existing = ws.cell(row=row, column=a_col).value
        if existing and str(existing).strip():
            print("   --> Already answered - skipping.")
            continue

        try:
            answer = pipeline.run(question)
        except Exception as e:
            answer = "ERROR: " + str(e)
            print("   [FAIL] " + answer)

        ws.cell(row=row, column=a_col, value=answer)
        answer_str = str(answer)
        preview = (answer_str[:120] + " ...") if len(answer_str) > 120 else answer_str
        print("   A: " + preview)

        # Save after every question so progress is not lost on crash
        wb.save(excel_path)

        if idx < total and delay > 0:
            print("   [WAIT] " + str(delay) + "s ...")
            time.sleep(delay)

    wb.save(excel_path)
    pipeline.close()

    print("\n" + "=" * 72)
    print("[DONE] Stress test complete.")
    print("       Results saved to: " + str(excel_path))
    print("=" * 72 + "\n")


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    THIS_DIR = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(description="GraphRAG Stress Test Runner")
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="Path to the .xlsx file. Defaults to the first .xlsx in this script's directory.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name. Defaults to the active (first) sheet.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=2.0,
        help="Seconds to wait between queries (avoids rate-limits). Default: 2.",
    )

    args = parser.parse_args()
    excel_file = args.excel if args.excel else find_excel(THIS_DIR)
    run_stress_test(excel_file, args.sheet, args.delay)
